#!/usr/bin/env python3
"""
NetSuite Material Composition Analyzer
=======================================
Analyzes material composition data from a NetSuite saved search CSV export.

Checks performed:
  1. Name suffix → material composition consistency
       -SS (or -SS-<anything>) → must be Stainless Steel
       -B (exact final segment) → must be Brass
       -BB is Brennan Black (a program/class), NOT Brass — not flagged
  2. Class name must reflect material composition
       e.g. "Hydraulic Fittings - 37° JIC - Steel" should contain "Steel"
       Brennan Black parts are exempt (material won't appear in that class by design)
  3. Matrix Material field must match Material Composition (when populated)
  4. Missing material composition
  5. Matrix parents detected (same Internal ID repeated → multiple compositions)

Sanity / data-quality guards baked in:
  - Column detection is case-insensitive and whitespace-tolerant
  - All string comparisons are case-insensitive and stripped
  - Encoding fallback (utf-8-sig → latin-1) for Excel-exported CSVs
  - -BB vs -B exact segment matching (no partial suffix collisions)
  - "Steel" in class does NOT satisfy "Stainless Steel" requirement
  - Multi-value composition fields (comma / semicolon / pipe separated)
  - Duplicate Internal ID detection → matrix parent flagging
  - Boolean flag columns clearly separated from informational notes
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime

# In PyInstaller --windowed builds on Windows, stdout/stderr are None.
# Redirect to devnull so print() calls don't crash; CLI usage from a real
# terminal is unaffected because stdout/stderr exist there.
if sys.stdout is None:
    sys.stdout = open(os.devnull, "w")
if sys.stderr is None:
    sys.stderr = open(os.devnull, "w")

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ── Constants ──────────────────────────────────────────────────────────────────

# Maps a normalized material composition string → keywords that MUST appear in
# the class name for the class to be considered consistent with the material.
MATERIAL_CLASS_KEYWORDS: dict[str, list[str]] = {
    "stainless steel":  ["stainless"],
    "steel":            ["steel"],          # "stainless" in class would be wrong — handled separately
    "brass":            ["brass"],
    "aluminum":         ["aluminum", "aluminium"],
    "aluminum alloy":   ["aluminum", "aluminium"],
    "carbon steel":     ["carbon"],
    "cast iron":        ["cast iron"],
    "copper":           ["copper"],
    "titanium":         ["titanium"],
    "plastic":          ["plastic", "nylon", "polymer"],
    "rubber":           ["rubber"],
    "zinc":             ["zinc"],
    "chrome":           ["chrome"],
    "nickel":           ["nickel"],
}

# Name suffix segment → expected Material Composition value.
# Keys are uppercase; values are the canonical display string.
SUFFIX_MATERIAL_MAP: dict[str, str] = {
    "SS": "Stainless Steel",
    "B":  "Brass",
}

# Segments that look like a suffix but are NOT material indicators.
# Prevents false-positive material detection.
NON_MATERIAL_SEGMENTS: set[str] = {
    "BB",   # Brennan Black program
    "LW",   # Light weight — structural modifier, not material
    "SPL",  # Special
    "D",    # Size/dimension indicator
    "UNP",  # Unpainted
    "ZN",   # Zinc plated — note: treated separately if needed
}

# Classes that are explicitly exempt from the "class must contain material" check.
EXEMPT_CLASS_SUBSTRINGS: list[str] = [
    "brennan black",
]

# Issue flags — these contribute to any_flag and warrant human review.
ISSUE_FLAGS: dict[str, str] = {
    "flag_suffix_mismatch":            "Name Suffix ↔ Material Mismatch",
    "flag_class_material_mismatch":    "Class Doesn't Reflect Material",
    "flag_material_field_mismatch":    "Matrix Material Field Mismatch",
    "flag_empty_material_composition": "Missing Material Composition",
}

# Informational flags — noted for context but do NOT trigger any_flag.
INFO_FLAGS: dict[str, str] = {
    "flag_no_suffix_non_steel": "No Material Suffix — Non-Steel Composition (Review)",
    "flag_is_matrix_parent":    "Matrix Parent (Multiple Compositions)",
    "flag_is_bb_part":          "Brennan Black Part (Class Check Exempt)",
}

# Combined ordered dict used for display/export.
FLAG_META: dict[str, str] = {**ISSUE_FLAGS, **INFO_FLAGS}
FLAG_COLS = list(FLAG_META.keys())


# ── Name Parsing ───────────────────────────────────────────────────────────────

def get_name_segments(name: str) -> list[str]:
    """Split an item name on '-' and return cleaned, uppercased segments."""
    if not name or str(name).strip() in ("", "nan", "None"):
        return []
    return [s.strip().upper() for s in str(name).split("-") if s.strip()]


def detect_suffix_material(name: str) -> str | None:
    """
    Infer expected material composition from name segments.

    Rules (applied in order):
      - Any segment == 'SS' → Stainless Steel
        (handles -SS-LW, -SS-SPL, etc. — SS anywhere still means stainless)
      - Final segment == 'B' exactly → Brass
        (-BB is NOT brass; exact match guards against this)

    Returns the canonical material string, or None if no suffix match.
    """
    segments = get_name_segments(name)
    if not segments:
        return None

    # SS can appear anywhere in the segment chain
    if "SS" in segments:
        return "Stainless Steel"

    # B must be the final segment and must be EXACTLY 'B'
    if segments[-1] == "B":
        return "Brass"

    return None


def is_bb_part(name: str, class_val: str) -> bool:
    """
    Return True if this part belongs to the Brennan Black program.
    BB parts are exempt from the class-contains-material check because
    the class IS 'Brennan Black' and won't contain the material by design.
    """
    segments = get_name_segments(name)
    class_lower = str(class_val).strip().lower() if class_val else ""

    if "BB" in segments:
        return True
    if any(sub in class_lower for sub in EXEMPT_CLASS_SUBSTRINGS):
        return True
    return False


# ── Material / Class Comparison ────────────────────────────────────────────────

def normalize(value: str) -> str:
    """Lowercase + strip a string; return '' for missing values."""
    if not value or str(value).strip() in ("", "nan", "None"):
        return ""
    return str(value).strip().lower()


def material_matches_class(material_comp: str, class_val: str) -> bool:
    """
    Return True if the class name is consistent with the material composition.

    Logic:
      1. Look up the material in MATERIAL_CLASS_KEYWORDS.
      2. At least one keyword must appear in the class string.
      3. Special guard: if material is plain 'steel' but class contains
         'stainless', that's a mismatch (wrong grade of steel).
      4. Unknown materials fall back to a simple substring check.
    """
    mat = normalize(material_comp)
    cls = normalize(class_val)

    if not mat or not cls:
        return True     # Cannot evaluate — skip, do not flag

    keywords = MATERIAL_CLASS_KEYWORDS.get(mat)

    if keywords is None:
        # Unknown material: simple substring check
        return mat in cls

    # Guard: plain "steel" should not accept a "stainless" class as a match
    if mat == "steel" and "stainless" in cls:
        return False

    return any(kw in cls for kw in keywords)


def split_multi_composition(value: str) -> list[str]:
    """
    Split a material composition field that may hold multiple values.
    Handles comma, semicolon, pipe, and newline delimiters.
    Returns a list of stripped, non-empty strings.
    """
    if not value or str(value).strip() in ("", "nan", "None"):
        return []
    parts = re.split(r"[,;\n|]+", str(value))
    return [p.strip() for p in parts if p.strip()]


# ── Core Analysis ──────────────────────────────────────────────────────────────

def _resolve_column(col_map: dict[str, str], *candidates: str) -> str | None:
    """Find the first matching column name (case-insensitive)."""
    for c in candidates:
        if c.lower() in col_map:
            return col_map[c.lower()]
    return None


def analyze_dataframe(
    df: pd.DataFrame,
    enabled_checks: set[str] | None = None,
) -> pd.DataFrame:
    """
    Main analysis entry point.
    Adds flag columns, 'expected_material_from_name', 'analysis_notes',
    and 'any_flag' to a copy of the input DataFrame.

    enabled_checks : set of ISSUE_FLAGS keys to evaluate.
                     None (default) → run all checks.
                     Disabled checks are still listed as columns but always False.
    """
    if enabled_checks is None:
        enabled_checks = set(ISSUE_FLAGS.keys())
    df = df.copy()

    # ── Normalize column names ─────────────────────────────────────────────
    df.columns = [c.strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}

    col_id       = _resolve_column(col_map, "internal id", "internalid", "internal_id", "id")
    col_name     = _resolve_column(col_map, "name", "item name", "item_name")
    col_class    = _resolve_column(col_map, "class", "item class", "item_class")
    col_material = _resolve_column(col_map, "material")
    col_mat_comp = _resolve_column(
        col_map, "material composition", "material_composition", "materialcomposition"
    )

    # Validate required columns
    missing = [
        label for label, col in [
            ("Name", col_name),
            ("Class", col_class),
            ("Material Composition", col_mat_comp),
        ]
        if col is None
    ]
    if missing:
        raise ValueError(
            f"Required column(s) not found: {', '.join(missing)}\n"
            f"Columns in file: {list(df.columns)}"
        )

    # ── Sanitize string fields ─────────────────────────────────────────────
    for col in filter(None, [col_id, col_name, col_class, col_material, col_mat_comp]):
        df[col] = df[col].astype(str).str.strip().replace(
            {"nan": "", "None": "", "NaN": "", "none": ""}
        )

    # ── Detect matrix parents (Internal ID appears more than once) ─────────
    matrix_parent_ids: set[str] = set()
    if col_id:
        counts = df[col_id].replace("", pd.NA).dropna().value_counts()
        matrix_parent_ids = set(counts[counts > 1].index)

    # ── Row-level analysis ─────────────────────────────────────────────────
    flag_rows: list[dict] = []

    for _, row in df.iterrows():
        name      = str(row[col_name]).strip()
        class_val = str(row[col_class]).strip()         if col_class    else ""
        mat_comp  = str(row[col_mat_comp]).strip()      if col_mat_comp else ""
        material  = str(row[col_material]).strip()      if col_material else ""
        int_id    = str(row[col_id]).strip()            if col_id       else ""

        # Replace sentinel strings left from astype(str)
        for sentinel in ("nan", "None", "NaN"):
            if mat_comp  == sentinel: mat_comp  = ""
            if material  == sentinel: material  = ""
            if class_val == sentinel: class_val = ""
            if name      == sentinel: name      = ""

        flags: dict[str, bool] = {f: False for f in FLAG_COLS}
        notes: list[str] = []

        expected_mat   = detect_suffix_material(name)
        # Default recommendation: suffix-derived material, or "Steel" for bare parts.
        recommended_mat = expected_mat if expected_mat else "Steel"
        bb              = is_bb_part(name, class_val)
        compositions    = split_multi_composition(mat_comp)

        flags["flag_is_bb_part"] = bb

        # Matrix parent
        if int_id and int_id in matrix_parent_ids:
            flags["flag_is_matrix_parent"] = True

        # ── Missing material composition ───────────────────────────────────
        if "flag_empty_material_composition" in enabled_checks and not mat_comp:
            flags["flag_empty_material_composition"] = True
            notes.append("Material Composition is blank")

        # ── Suffix ↔ material mismatch (bidirectional) ─────────────────────
        if "flag_suffix_mismatch" in enabled_checks and mat_comp:
            segments = get_name_segments(name)
            mat_norm = normalize(mat_comp)

            # Forward: name suffix implies a material that disagrees with composition
            if expected_mat and normalize(expected_mat) != mat_norm:
                flags["flag_suffix_mismatch"] = True
                notes.append(
                    f"Name suffix implies '{expected_mat}' "
                    f"but Material Composition is '{mat_comp}'"
                )

            # Reverse: composition is a suffix-required material but name has no
            # such suffix. Skip multi-composition rows (matrix parents) where
            # multiple materials legitimately coexist on one row.
            elif len(compositions) == 1 and not flags["flag_is_matrix_parent"]:
                if mat_norm == "stainless steel" and "SS" not in segments:
                    flags["flag_suffix_mismatch"] = True
                    notes.append(
                        "Material Composition is 'Stainless Steel' but name has "
                        "no -SS suffix (likely should be Steel)"
                    )
                elif mat_norm == "brass" and (not segments or segments[-1] != "B"):
                    flags["flag_suffix_mismatch"] = True
                    notes.append(
                        "Material Composition is 'Brass' but name does not end "
                        "in -B (likely should be Steel)"
                    )

        # ── Class does not reflect material ────────────────────────────────
        if "flag_class_material_mismatch" in enabled_checks and mat_comp and not bb:
            class_mismatches = [
                comp for comp in compositions
                if not material_matches_class(comp, class_val)
            ]
            if class_mismatches:
                flags["flag_class_material_mismatch"] = True
                notes.append(
                    f"Class '{class_val}' does not reflect: "
                    + ", ".join(f"'{c}'" for c in class_mismatches)
                )

        # ── Matrix Material field mismatch ─────────────────────────────────
        # Only evaluated when the Material field is populated (matrix part).
        if "flag_material_field_mismatch" in enabled_checks and material and mat_comp:
            comp_normalized = [normalize(c) for c in compositions]
            if normalize(material) not in comp_normalized:
                flags["flag_material_field_mismatch"] = True
                notes.append(
                    f"Matrix Material '{material}' not in "
                    f"Material Composition '{mat_comp}'"
                )

        # ── No suffix → non-Steel composition (soft recommendation) ────────
        # Most suffix-less parts are plain Steel. Flag for review when they
        # have a different composition so a human can confirm it's intentional.
        # Suppressed when flag_suffix_mismatch already fires for the same row
        # (avoid double-flagging the SS/Brass cases now caught by the reverse rule).
        if (not expected_mat and mat_comp
                and normalize(mat_comp) != "steel"
                and not flags["flag_suffix_mismatch"]):
            flags["flag_no_suffix_non_steel"] = True
            notes.append(
                f"No material suffix — composition is '{mat_comp}' "
                f"(most bare parts are Steel; confirm if intentional)"
            )

        flag_rows.append({
            **flags,
            "recommended_material":        recommended_mat,
            "expected_material_from_name": expected_mat or "",
            "analysis_notes":              " | ".join(notes) if notes else "OK",
            "any_flag":                    any(flags[k] for k in ISSUE_FLAGS),
        })

    result_df = pd.DataFrame(flag_rows, index=df.index)
    return pd.concat([df, result_df], axis=1)


# ── Terminal Report ────────────────────────────────────────────────────────────

def print_summary(df: pd.DataFrame) -> None:
    total   = len(df)
    flagged = int(df["any_flag"].sum())
    pct     = flagged / total * 100 if total else 0.0

    width = 62
    sep   = "═" * width

    print(f"\n{sep}")
    print("  NETSUITE MATERIAL COMPOSITION — ANALYSIS REPORT")
    print(sep)
    print(f"  {'Total records analyzed':<40} {total:>8,}")
    print(f"  {'Records with issues':<40} {flagged:>8,}  ({pct:.1f}%)")
    print(f"  {'Clean records':<40} {total - flagged:>8,}")
    print()
    print(f"  {'── Issue Breakdown':─<{width - 4}}")
    for col, label in FLAG_META.items():
        if col in df.columns:
            count = int(df[col].sum())
            marker = "⚠" if count > 0 else " "
            print(f"  {marker} {label:<42} {count:>6,}")
    print(sep)


def print_flagged_details(df: pd.DataFrame, col_name: str) -> None:
    flagged = df[df["any_flag"] == True]
    if flagged.empty:
        print("\nNo flagged records.")
        return

    print(f"\nFlagged Records ({len(flagged):,}):\n")
    cols = [c for c in [col_name, "analysis_notes"] if c in flagged.columns]
    print(flagged[cols].to_string(index=False, max_colwidth=80))


# ── Excel Export ───────────────────────────────────────────────────────────────

# Color palette
_C = {
    "header_bg":   "1F3864",
    "header_fg":   "FFFFFF",
    "flag_red":    "FFB3B3",
    "flag_yellow": "FFF3CC",
    "ok_green":    "D6F0D6",
    "row_alt":     "F2F4F8",
    "row_plain":   "FFFFFF",
    "blue_accent": "CCE5FF",
    "border":      "BFBFBF",
}

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border() -> "Border":
    side = Side(style="thin", color=_C["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def export_excel(df: pd.DataFrame, output_path: str) -> None:
    if not EXCEL_AVAILABLE:
        print(
            "\nERROR: openpyxl is not installed.\n"
            "Install it with:  pip install openpyxl\n"
        )
        sys.exit(1)

    # Column ordering for data sheets: source columns first, then derived
    source_cols  = [c for c in df.columns if c not in FLAG_COLS + [
        "any_flag", "expected_material_from_name", "analysis_notes"
    ]]
    derived_cols = ["recommended_material", "expected_material_from_name", "analysis_notes"]
    flag_cols    = [c for c in FLAG_COLS if c in df.columns] + ["any_flag"]
    ordered_cols = source_cols + derived_cols + flag_cols

    wb = Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    _build_summary_sheet(wb.active, df)

    # ── Sheet 2: All Data ──────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Data")
    _build_data_sheet(ws_all, df[ordered_cols])

    # ── Sheet 3: Flagged Items ─────────────────────────────────────────────
    ws_flag = wb.create_sheet("Flagged Items")
    _build_data_sheet(ws_flag, df[df["any_flag"] == True][ordered_cols])

    # ── Per-flag sheets (only when issues exist) ───────────────────────────
    for flag_col, label in FLAG_META.items():
        if flag_col in df.columns and df[flag_col].sum() > 0:
            sheet_name = label[:31]   # Excel sheet name limit
            ws = wb.create_sheet(sheet_name)
            _build_data_sheet(ws, df[df[flag_col] == True][ordered_cols])

    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")


def _build_summary_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14

    # Title banner
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "NetSuite Material Composition Analysis"
    cell.font  = Font(bold=True, size=15, color=_C["header_fg"])
    cell.fill  = _fill(_C["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Timestamp
    ws.merge_cells("A2:B2")
    ts = ws["A2"]
    ts.value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    ts.font  = Font(italic=True, size=10, color="555555")
    ts.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.append([])  # spacer

    total   = len(df)
    flagged = int(df["any_flag"].sum())

    summary_rows = [
        ("Total Records Analyzed",  f"{total:,}"),
        ("Records with Issues",     f"{flagged:,}"),
        ("Clean Records",           f"{total - flagged:,}"),
        ("Issue Rate",              f"{flagged / total * 100:.1f}%" if total else "—"),
        ("", ""),
        ("Issue Breakdown",         "Count"),
    ]
    for flag_col, label in FLAG_META.items():
        count = int(df[flag_col].sum()) if flag_col in df.columns else 0
        summary_rows.append((label, f"{count:,}"))

    for row_data in summary_rows:
        ws.append(list(row_data))
        row_idx = ws.max_row
        a_cell = ws.cell(row=row_idx, column=1)
        b_cell = ws.cell(row=row_idx, column=2)

        for cell in (a_cell, b_cell):
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")

        if row_data[0] == "Issue Breakdown":
            for cell in (a_cell, b_cell):
                cell.fill = _fill(_C["blue_accent"])
                cell.font = Font(bold=True)
        elif row_data[0] in ("Total Records Analyzed", "Records with Issues",
                              "Clean Records", "Issue Rate"):
            a_cell.font = Font(bold=True)

        ws.row_dimensions[row_idx].height = 18


def _build_data_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet with professional formatting."""
    FRIENDLY: dict[str, str] = {
        "flag_suffix_mismatch":            "Suffix Mismatch",
        "flag_class_material_mismatch":    "Class Mismatch",
        "flag_material_field_mismatch":    "Matrix Field Mismatch",
        "flag_empty_material_composition": "Missing Composition",
        "flag_is_matrix_parent":           "Matrix Parent",
        "flag_is_bb_part":                 "BB Part (Exempt)",
        "any_flag":                        "Has Issue",
        "recommended_material":            "Recommended Material",
        "expected_material_from_name":     "Suffix-Detected Material",
        "analysis_notes":                  "Analysis Notes",
    }

    headers = [FRIENDLY.get(c, c) for c in df.columns]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _fill(_C["header_bg"])
        cell.font      = Font(bold=True, color=_C["header_fg"], size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Identify key columns by position
    col_names = list(df.columns)
    any_flag_pos = col_names.index("any_flag") + 1 if "any_flag" in col_names else None
    notes_pos    = col_names.index("analysis_notes") + 1 if "analysis_notes" in col_names else None

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        values = list(row)
        ws.append(values)

        has_issue = bool(values[any_flag_pos - 1]) if any_flag_pos else False
        row_bg    = _C["flag_red"] if has_issue else (
            _C["row_alt"] if row_idx % 2 == 0 else _C["row_plain"]
        )

        for col_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = _border()
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == notes_pos)
            )

            if col_name.startswith("flag_") or col_name == "any_flag":
                val = cell.value
                if val is True or str(val).upper() == "TRUE":
                    cell.value = "YES"
                    cell.fill  = _fill(_C["flag_red"])
                    cell.font  = Font(bold=True, color="990000")
                else:
                    cell.value = "—"
                    cell.fill  = _fill(_C["row_plain"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = _fill(row_bg)

        ws.row_dimensions[row_idx].height = 15

    # Column widths
    for col_idx, col_name in enumerate(col_names, 1):
        letter = get_column_letter(col_idx)
        if col_name == "analysis_notes":
            ws.column_dimensions[letter].width = 52
        elif col_name.startswith("flag_") or col_name == "any_flag":
            ws.column_dimensions[letter].width = 16
        elif col_name in ("recommended_material", "expected_material_from_name"):
            ws.column_dimensions[letter].width = 26
        else:
            max_content = max(
                (len(str(v)) for v in df[col_name] if v is not None),
                default=0
            )
            header_len = len(FRIENDLY.get(col_name, col_name))
            ws.column_dimensions[letter].width = min(max(max_content, header_len) + 3, 42)

    # Auto-filter
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


# ── CLI ────────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="analyze_materials",
        description="Analyze NetSuite material composition data from a CSV saved-search export.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python analyze_materials.py items.csv
  python analyze_materials.py items.csv --excel report.xlsx
  python analyze_materials.py items.csv --excel report.xlsx --flagged-only
        """,
    )
    p.add_argument("input",
        help="Path to the CSV file exported from NetSuite"
    )
    p.add_argument("--excel", "-e", metavar="OUTPUT.xlsx",
        help="Also export results to a formatted Excel workbook"
    )
    p.add_argument("--flagged-only", "-f", action="store_true",
        help="Print only flagged records in the terminal output"
    )
    p.add_argument("--encoding", default="utf-8-sig",
        help="CSV file encoding (default: utf-8-sig — handles Excel BOM automatically)"
    )
    p.add_argument(
        "--checks", "-c",
        nargs="+",
        choices=["suffix", "class", "matrix", "empty", "all"],
        default=["all"],
        help=(
            "Which checks to run (default: all). "
            "'suffix' = name ↔ material; 'class' = class reflects material; "
            "'matrix' = matrix Material field; 'empty' = missing composition. "
            "Example: --checks suffix"
        ),
    )
    return p


# Maps user-friendly check names → the underlying flag column.
CHECK_NAME_TO_FLAG: dict[str, str] = {
    "suffix": "flag_suffix_mismatch",
    "class":  "flag_class_material_mismatch",
    "matrix": "flag_material_field_mismatch",
    "empty":  "flag_empty_material_composition",
}


def resolve_checks(names: list[str]) -> set[str]:
    """Convert a list like ['suffix','class'] or ['all'] → set of flag column keys."""
    if not names or "all" in names:
        return set(ISSUE_FLAGS.keys())
    return {CHECK_NAME_TO_FLAG[n] for n in names if n in CHECK_NAME_TO_FLAG}


def load_csv(path: Path, encoding: str) -> pd.DataFrame:
    """Load CSV with automatic encoding fallback."""
    try:
        df = pd.read_csv(path, encoding=encoding, dtype=str)
        print(f"Loaded {len(df):,} rows from {path}")
        return df
    except UnicodeDecodeError:
        print(f"  utf-8 decode failed — retrying with latin-1...")
        df = pd.read_csv(path, encoding="latin-1", dtype=str)
        print(f"Loaded {len(df):,} rows from {path} (latin-1)")
        return df


def run_cli(args) -> None:
    csv_path = Path(args.input)
    if not csv_path.exists():
        print(f"ERROR: File not found: {csv_path}")
        sys.exit(1)

    df_raw = load_csv(csv_path, args.encoding)
    print(f"Columns: {list(df_raw.columns)}\n")

    try:
        df_result = analyze_dataframe(df_raw, enabled_checks=resolve_checks(args.checks))
    except ValueError as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)

    col_map  = {c.lower(): c for c in df_result.columns}
    col_name = next(
        (col_map[k] for k in ("name", "item name", "item_name") if k in col_map),
        df_result.columns[0]
    )

    print_summary(df_result)

    if args.flagged_only:
        print_flagged_details(df_result, col_name)
    else:
        flagged_count = int(df_result["any_flag"].sum())
        if flagged_count:
            print_flagged_details(df_result, col_name)

    if args.excel:
        export_excel(df_result, args.excel)


# ── GUI ────────────────────────────────────────────────────────────────────────

def run_gui() -> None:
    """Simple tkinter GUI: pick CSV, click Analyze, get an Excel report."""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import threading
    import subprocess
    import os

    root = tk.Tk()
    root.title("Material Composition Analyzer")
    root.geometry("620x480")
    root.minsize(540, 420)

    state = {"csv_path": None, "xlsx_path": None}

    # ── Layout ─────────────────────────────────────────────────────────────
    main_frame = ttk.Frame(root, padding=20)
    main_frame.pack(fill="both", expand=True)

    title_lbl = ttk.Label(
        main_frame,
        text="NetSuite Material Composition Analyzer",
        font=("Segoe UI", 14, "bold"),
    )
    title_lbl.pack(anchor="w")

    subtitle_lbl = ttk.Label(
        main_frame,
        text="Select a CSV exported from your NetSuite saved search, then click Analyze.",
        foreground="#555555",
    )
    subtitle_lbl.pack(anchor="w", pady=(2, 16))

    # File picker row
    file_frame = ttk.LabelFrame(main_frame, text="Input CSV", padding=10)
    file_frame.pack(fill="x")

    file_lbl = ttk.Label(file_frame, text="(no file selected)", foreground="#888888")
    file_lbl.pack(side="left", fill="x", expand=True)

    def pick_file():
        path = filedialog.askopenfilename(
            title="Select NetSuite CSV export",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            state["csv_path"] = path
            file_lbl.config(text=Path(path).name, foreground="#000000")
            analyze_btn.config(state="normal")
            status_lbl.config(text="Ready to analyze.", foreground="#000000")

    pick_btn = ttk.Button(file_frame, text="Choose CSV…", command=pick_file)
    pick_btn.pack(side="right", padx=(10, 0))

    # ── Checks scope ───────────────────────────────────────────────────────
    checks_frame = ttk.LabelFrame(main_frame, text="Checks to Run", padding=10)
    checks_frame.pack(fill="x", pady=(12, 0))

    check_vars: dict[str, tk.BooleanVar] = {
        "suffix": tk.BooleanVar(value=True),
        "class":  tk.BooleanVar(value=True),
        "matrix": tk.BooleanVar(value=True),
        "empty":  tk.BooleanVar(value=True),
    }
    check_labels = {
        "suffix": "Name suffix ↔ material composition",
        "class":  "Class reflects material composition",
        "matrix": "Matrix Material field matches composition",
        "empty":  "Missing material composition",
    }

    for key, label in check_labels.items():
        ttk.Checkbutton(checks_frame, text=label, variable=check_vars[key]).pack(
            anchor="w"
        )

    # Quick scope buttons
    scope_btns = ttk.Frame(checks_frame)
    scope_btns.pack(anchor="w", pady=(8, 0))

    def set_scope(*keys):
        for k, var in check_vars.items():
            var.set(k in keys)

    ttk.Button(scope_btns, text="All",
               command=lambda: set_scope("suffix", "class", "matrix", "empty"),
               width=10).pack(side="left")
    ttk.Button(scope_btns, text="Suffix only",
               command=lambda: set_scope("suffix"),
               width=12).pack(side="left", padx=(6, 0))
    ttk.Button(scope_btns, text="Class only",
               command=lambda: set_scope("class"),
               width=12).pack(side="left", padx=(6, 0))

    # Status / summary area
    status_frame = ttk.LabelFrame(main_frame, text="Status", padding=10)
    status_frame.pack(fill="both", expand=True, pady=(12, 0))

    status_lbl = ttk.Label(
        status_frame,
        text="Choose a CSV file to begin.",
        foreground="#555555",
    )
    status_lbl.pack(anchor="w")

    summary_text = tk.Text(
        status_frame,
        height=10,
        wrap="word",
        state="disabled",
        font=("Consolas", 10),
        background="#f7f7f7",
        relief="flat",
    )
    summary_text.pack(fill="both", expand=True, pady=(8, 0))

    progress = ttk.Progressbar(main_frame, mode="indeterminate")

    # Action buttons
    btn_frame = ttk.Frame(main_frame)
    btn_frame.pack(fill="x", pady=(12, 0))

    def open_path(path: str):
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path], check=False)
            else:
                subprocess.run(["xdg-open", path], check=False)
        except Exception as exc:
            messagebox.showerror("Couldn't open", str(exc))

    open_file_btn = ttk.Button(
        btn_frame,
        text="Open Excel Report",
        command=lambda: open_path(state["xlsx_path"]) if state["xlsx_path"] else None,
        state="disabled",
    )
    open_folder_btn = ttk.Button(
        btn_frame,
        text="Open Folder",
        command=lambda: open_path(str(Path(state["xlsx_path"]).parent)) if state["xlsx_path"] else None,
        state="disabled",
    )

    def set_summary(text: str):
        summary_text.config(state="normal")
        summary_text.delete("1.0", "end")
        summary_text.insert("1.0", text)
        summary_text.config(state="disabled")

    def do_analyze():
        csv_path = state["csv_path"]
        if not csv_path:
            return

        analyze_btn.config(state="disabled")
        pick_btn.config(state="disabled")
        open_file_btn.config(state="disabled")
        open_folder_btn.config(state="disabled")
        progress.pack(fill="x", pady=(10, 0))
        progress.start(10)
        status_lbl.config(text="Analyzing… please wait.", foreground="#000000")
        set_summary("")

        # Snapshot the user's check selection at click time
        selected = [k for k, v in check_vars.items() if v.get()]
        if not selected:
            messagebox.showwarning(
                "No checks selected",
                "Pick at least one check to run."
            )
            analyze_btn.config(state="normal")
            pick_btn.config(state="normal")
            return
        enabled = resolve_checks(selected)

        # Filename suffix reflecting scope: "all" or e.g. "suffix" or "suffix_class"
        is_all = (set(selected) == set(check_vars.keys()))
        scope_tag = "all" if is_all else "_".join(selected)

        def worker():
            try:
                csv_p = Path(csv_path)
                df_raw = load_csv(csv_p, "utf-8-sig")
                df_result = analyze_dataframe(df_raw, enabled_checks=enabled)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                xlsx_p = csv_p.with_name(
                    f"{csv_p.stem}_analysis_{scope_tag}_{timestamp}.xlsx"
                )
                export_excel(df_result, str(xlsx_p))

                total   = len(df_result)
                flagged = int(df_result["any_flag"].sum())
                pct     = flagged / total * 100 if total else 0.0

                lines = [
                    f"Analyzed:       {total:,} records",
                    f"With issues:    {flagged:,}  ({pct:.1f}%)",
                    f"Clean:          {total - flagged:,}",
                    "",
                    "Issue breakdown:",
                ]
                for col, label in FLAG_META.items():
                    if col in df_result.columns:
                        cnt = int(df_result[col].sum())
                        marker = "  !" if cnt > 0 else "   "
                        lines.append(f"{marker} {label:<48} {cnt:>6,}")
                lines += ["", f"Saved to:  {xlsx_p.name}"]

                state["xlsx_path"] = str(xlsx_p)

                def on_done():
                    progress.stop()
                    progress.pack_forget()
                    set_summary("\n".join(lines))
                    status_lbl.config(text="Done.", foreground="#006600")
                    analyze_btn.config(state="normal")
                    pick_btn.config(state="normal")
                    open_file_btn.config(state="normal")
                    open_folder_btn.config(state="normal")
                root.after(0, on_done)

            except Exception as exc:
                err_msg = str(exc)
                def on_err():
                    progress.stop()
                    progress.pack_forget()
                    status_lbl.config(text="Error.", foreground="#990000")
                    set_summary(f"ERROR:\n\n{err_msg}")
                    analyze_btn.config(state="normal")
                    pick_btn.config(state="normal")
                root.after(0, on_err)

        threading.Thread(target=worker, daemon=True).start()

    analyze_btn = ttk.Button(
        btn_frame,
        text="Analyze and Save Excel",
        command=do_analyze,
        state="disabled",
    )
    analyze_btn.pack(side="left")
    open_file_btn.pack(side="left", padx=(8, 0))
    open_folder_btn.pack(side="left", padx=(8, 0))

    quit_btn = ttk.Button(btn_frame, text="Close", command=root.destroy)
    quit_btn.pack(side="right")

    root.mainloop()


def main() -> None:
    # No CLI args → launch GUI (this is what double-clicking the .exe does).
    if len(sys.argv) <= 1:
        run_gui()
        return

    # Otherwise, parse CLI args as before.
    args = build_parser().parse_args()
    run_cli(args)


if __name__ == "__main__":
    main()
